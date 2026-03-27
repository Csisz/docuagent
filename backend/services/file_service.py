"""
Fájlkezelő service.
Szöveg kinyerése különböző formátumokból + nyelvdetektálás.
"""
import logging
from pathlib import Path

log = logging.getLogger("docuagent")


def extract_text(path: Path) -> str:
    """Szöveg kinyerése PDF, DOCX, XLSX, és szöveges fájlokból."""
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(path)
        elif ext in (".docx", ".doc"):
            return _extract_docx(path)
        elif ext in (".xlsx", ".xls"):
            return _extract_xlsx(path)
        else:
            return path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        log.warning(f"Text extraction failed for {path.name}: {e}")
        return ""


def detect_language(text: str) -> str:
    """Egyszerű nyelvdetektálás szógyakoriság alapján. Visszaad: HU / DE / EN."""
    t = text.lower()
    hu = sum(1 for w in ["és", "a", "az", "hogy", "van", "nem"] if f" {w} " in t)
    de = sum(1 for w in ["und", "die", "der", "das", "ist", "nicht"] if f" {w} " in t)
    if hu > de:
        return "HU"
    elif de > 0:
        return "DE"
    return "EN"


# ── Belső függvények ──────────────────────────────────────────

def _extract_pdf(path: Path) -> str:
    try:
        import fitz
        doc = fitz.open(str(path))
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        if text.strip():
            return text
    except ImportError:
        pass
    import pdfminer.high_level as pm
    return pm.extract_text(str(path))


def _extract_docx(path: Path) -> str:
    import docx
    d = docx.Document(str(path))
    return "\n".join(p.text for p in d.paragraphs if p.text.strip())


def _extract_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            r = " | ".join(str(c) for c in row if c is not None)
            if r.strip():
                parts.append(r)
    wb.close()
    return "\n".join(parts)
